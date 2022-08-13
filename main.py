import easyocr
import os
import re
import cv2
import colorsystem
import openpyxl as op
import pymysql

db = pymysql.connect(host='localhost', user='root', passwd='root', port=3306, db='hesuan_result_collection')
cursor = db.cursor()

student_sno = []
timeres = []
timeres_re = []  # 时间检测结果
testres = []  # 核酸检测结果
str = ''
pattern1 = r'采集时间[\d,\-,:,.]*'
f = open('result.txt', 'w')
f.truncate()
sql_delete = "Update xinan set time_result = '' , test_result = '' "
cursor.execute(sql_delete)
db.commit()

directory_name = r"D:\pycharm\OCR\testimg2"  # 核酸检测图片——文件夹路径
sql_count = "select count(*) from xinan"
cursor.execute(sql_count)
student_num = cursor.fetchone()[0]  # 班级总人数
student_checked = 0
student_failed = 0
student_left = student_num
f = open('result.txt', 'w')
for filename in os.listdir(directory_name):
    reader = easyocr.Reader(['ch_sim', 'en'], gpu=False)  # GPU or CPU
    result = reader.readtext(directory_name + r'/' + filename, detail=0)
    result = str.join(result)
    result = result.replace(" ", '')
    result = result.replace("|", '1')  # 去除杂质
    # print(result)
    f.write(result)
    f.write("\n")
    filename_stu = filename.replace(".jpg", "")
    student_sno.append(filename_stu)
    picture_flag = 1
    if picture_flag:
        try:
            timeresult = re.search('采集时间[\d,\-,:,.]*', result)
            timeres.append(timeresult.group()[4:])

        except Exception as err:
            timeres.append("时间匹配失败")
        try:
            frame = cv2.imread(directory_name + r'/' + filename)
            testresult = colorsystem.get_color(frame)  # 识别图片颜色
            if testresult == 'colorList/green':
                testresult = "阴性"
            elif testresult == 'colorList/red':
                testresult = "阳性"
            testres.append(testresult)
        except Exception as err:
            testres.append("结果存疑")
    else:
        timeres.append("图片错误")
        testres.append("图片错误")

for x in timeres:
    try:
        xx = list(x)
        xx.insert(10, ' ')  # 在新列表xx中插入空格
        xx[13] = ':'  # 将xx中的所有“:”和“.”都不加识别地替换为“:”
        xx[16] = ':'  # 将xx中的所有“:”和“.”都不加识别地替换为“:”
        xxx = ''.join(xx)  # 将xx列表中的各个元素合并，形成新的字符串记作xxx，print(xxx) = [2022-05-2316:22:18]
        timeres_re.append(xxx)
        student_checked = student_checked + 1
    except Exception as err:
        timeres_re.append("时间匹配失败")

timeres = timeres_re
print(student_sno)
print(timeres)
print(testres)

bg = op.load_workbook(r"result.xlsx")  # 应先将excel文件放入到工作目录下
bg.remove(bg["Sheet1"])
bg.create_sheet("Sheet1", index=0)
sheet = bg["Sheet1"]
sheet.cell(1, 1, "学号")
sheet.cell(1, 2, "姓名")
sheet.cell(1, 3, "检测时间")
sheet.cell(1, 4, "检测结果")

if len(timeres) == len(testres):
    for i in range(1, len(timeres) + 1):
        sql_fetch_name = "select sname from xinan where sno = '%s' " % student_sno[i - 1]
        cursor.execute(sql_fetch_name)

        sheet.cell(i + 1, 1, student_sno[i - 1])
        try:
            sheet.cell(i + 1, 2, cursor.fetchone()[0])
        except Exception as err:
            sheet.cell(i + 1, 2, "找不到姓名")
        sheet.cell(i + 1, 3, timeres[i - 1])
        sheet.cell(i + 1, 4, testres[i - 1])
        bg.save(r"result.xlsx")  # 对文件进行保存
        sql_update = "UPDATE `xinan` " \
                     "SET `time_result` ='%s' , `test_result` = '%s' " \
                     "WHERE sno = '%s'; " % (timeres[i - 1], testres[i - 1], student_sno[i - 1])
        try:
            sql_update_result = cursor.execute(sql_update)
            db.commit()
        except Exception as err:
            print("数据库写入失败:", err)
else:
    print("长度不匹配")
print("finished")
